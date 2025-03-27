from PySide6.QtWidgets import (QWidget, QVBoxLayout, QPushButton, 
                             QTextEdit, QFileDialog, QMessageBox,
                             QHBoxLayout, QTableWidget, QTableWidgetItem,
                             QHeaderView)
from PySide6.QtCore import Qt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class FilePreviewWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("文件预览")
        self.setMinimumSize(1000, 800)
        
        # 创建主布局
        layout = QVBoxLayout()
        
        # 创建按钮布局
        button_layout = QHBoxLayout()
        
        # 创建选择文件按钮
        self.select_button = QPushButton("选择Excel文件")
        self.select_button.clicked.connect(self.select_file)
        button_layout.addWidget(self.select_button)
        
        # 创建规范检查按钮
        self.check_button = QPushButton("规范检查")
        self.check_button.clicked.connect(self.check_file)
        self.check_button.setEnabled(False)  # 初始状态禁用
        button_layout.addWidget(self.check_button)
        
        layout.addLayout(button_layout)
        
        # 创建文本显示区域
        self.text_area = QTextEdit()
        self.text_area.setReadOnly(True)
        layout.addWidget(self.text_area)
        
        # 创建表格显示区域
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["行号", "B列内容", "D列内容", "F列内容"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
        
        # 初始化变量
        self.current_file = None
        
    def select_file(self):
        """选择Excel文件并显示内容"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择Excel文件",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            try:
                # 读取Excel文件
                self.current_file = file_path
                df = pd.read_excel(file_path)
                
                # 显示文件信息
                info_text = f"文件路径: {file_path}\n"
                info_text += f"表格数量: {len(pd.ExcelFile(file_path).sheet_names)}\n"
                info_text += f"行数: {len(df)}\n"
                info_text += f"列数: {len(df.columns)}\n\n"
                
                try:
                    # 获取数据（跳过前5行标题）
                    data = df.iloc[5:]
                    
                    # 获取日期范围（A列）
                    dates = pd.to_datetime(data.iloc[:, 0], errors='coerce')
                    date_range = f"{dates.min().strftime('%m.%d')}-{dates.max().strftime('%m.%d')}"
                    
                    # 统计总项数（E列非空值且非nan）
                    construction_units = data.iloc[:, 4].astype(str)
                    valid_units = construction_units[construction_units.str.strip() != 'nan']
                    total_items = len(valid_units)
                    
                    # 统计风险等级（K列）
                    risk_counts = data.iloc[:, 10].value_counts()
                    low_risk = risk_counts.get('低风险', 0)
                    acceptable = risk_counts.get('可接受', 0)
                    
                    # 统计本单位和外施工单位（E列）
                    internal_units = ['计量电网运维班', '计量用户运维一班', '计量用户运维二班']
                    internal_mask = valid_units.isin(internal_units)
                    internal_count = internal_mask.sum()
                    external_count = (~internal_mask).sum()
                    
                    # 统计已发布项数（O列包含"已在系统发布"）
                    published_count = data.iloc[:, 14].astype(str).str.contains('已在系统发布', na=False).sum()
                    
                    # 生成统计信息
                    summary = f"（{date_range}）作业计划共审批{total_items}项（其中，低风险{low_risk}项，可接受{acceptable}项；本单位{internal_count}项，外施工单位{external_count}项），已在系统发布{published_count}项。\n\n"
                    info_text += summary
                    
                    # 显示施工单位详细统计
                    info_text += "施工单位统计:\n"
                    unit_counts = valid_units.value_counts()
                    for unit, count in unit_counts.items():
                        if unit.strip():  # 排除空字符串
                            info_text += f"{unit}: {count}条\n"
                            
                except Exception as e:
                    info_text += f"统计过程中出错：{str(e)}\n"
                
                self.text_area.setText(info_text)
                self.check_button.setEnabled(True)  # 启用规范检查按钮
                
            except Exception as e:
                QMessageBox.critical(self, "错误", f"读取文件时出错：\n{str(e)}")
                self.text_area.setText("")
                self.current_file = None
                self.check_button.setEnabled(False)  # 禁用规范检查按钮
                
    def check_file(self):
        """检查文件规范"""
        if not self.current_file:
            QMessageBox.warning(self, "警告", "请先选择Excel文件！")
            return
            
        try:
            # 加载工作簿
            wb = load_workbook(self.current_file)
            ws = wb.active
            
            # 清空表格
            self.table.setRowCount(0)
            
            # 创建黄色填充样式
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # 存储不规范的行
            invalid_rows = []
            
            # 从第7行开始检查
            for row in range(7, ws.max_row + 1):
                # 检查E列是否为指定内容
                e_value = str(ws[f'E{row}'].value).strip()
                if e_value in ["计量用户运维一班", "计量用户运维二班"]:
                    # 获取B、D、F列的值
                    b_value = str(ws[f'B{row}'].value).strip()
                    d_value = str(ws[f'D{row}'].value).strip()
                    f_value = str(ws[f'F{row}'].value).strip()
                    
                    # 检查B列和F列是否包含D列的内容
                    if d_value not in b_value or d_value not in f_value:
                        # 标记不规范的单元格
                        if d_value not in b_value:
                            ws[f'B{row}'].fill = yellow_fill
                        if d_value not in f_value:
                            ws[f'F{row}'].fill = yellow_fill
                            
                        # 添加到不规范行列表
                        invalid_rows.append({
                            'row': row,
                            'b': b_value,
                            'd': d_value,
                            'f': f_value,
                            'type': '内容不匹配'
                        })
                
                # 检查K列和N列
                k_value = str(ws[f'K{row}'].value).strip()
                n_value = str(ws[f'N{row}'].value).strip()
                
                if k_value == "可接受" and n_value != "否":
                    ws[f'N{row}'].fill = yellow_fill
                    invalid_rows.append({
                        'row': row,
                        'b': '可接受',
                        'd': n_value,
                        'f': 'N列应为"否"',
                        'type': '可接受风险'
                    })
                elif k_value == "低风险" and n_value != "是":
                    ws[f'N{row}'].fill = yellow_fill
                    invalid_rows.append({
                        'row': row,
                        'b': '低风险',
                        'd': n_value,
                        'f': 'N列应为"是"',
                        'type': '低风险'
                    })
            
            # 保存修改后的文件
            wb.save(self.current_file)
            
            # 显示不规范的行
            self.table.setRowCount(len(invalid_rows))
            for i, row_data in enumerate(invalid_rows):
                self.table.setItem(i, 0, QTableWidgetItem(str(row_data['row'])))
                self.table.setItem(i, 1, QTableWidgetItem(row_data['b']))
                self.table.setItem(i, 2, QTableWidgetItem(row_data['d']))
                self.table.setItem(i, 3, QTableWidgetItem(row_data['f']))
            
            # 显示检查结果
            result_text = f"\n检查完成！\n"
            result_text += f"共发现 {len(invalid_rows)} 处不规范内容\n"
            result_text += f"不规范内容已用黄色标记，详细信息见下方表格"
            
            self.text_area.append(result_text)
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"检查文件时出错：\n{str(e)}")
                
    def get_current_file(self):
        """获取当前选择的文件路径"""
        return self.current_file 