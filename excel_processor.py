import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from datetime import datetime
import os

class ExcelProcessor:
    REQUIRED_COLUMNS = [
        "序号", "作业类型（内容）", "项目管理单位/部门", "供电所", "施工单位",
        "施工地点", "工作开始时间", "工作结束时间", "工作负责人及电话（电话可选填）",
        "专业", "基准风险等级", "是否需要停电", "施工人数", "是否纳入视频监督", "备注"
    ]

    def __init__(self):
        self.duplicate_rows = []
        self.merged_data = None

    def validate_headers_df(self, df):
        """验证数据框的表头结构，返回(是否有效, 错误信息)"""
        try:
            # 获取表头列名
            columns = df.columns.tolist()
            
            # 检查B列是否为"作业类型（内容）"
            if len(columns) <= 1:
                return False, "表格结构错误：表格列数不足，至少需要15列"
            if columns[1] != "作业类型（内容）":
                return False, f"表头错误：B列的表头应为'作业类型（内容）'，当前为'{columns[1]}'"
            
            # 检查必需的列是否都存在
            missing_columns = []
            for col in self.REQUIRED_COLUMNS:
                if col not in columns:
                    missing_columns.append(col)
            
            if missing_columns:
                return False, f"缺少必要列：{', '.join(missing_columns)}"
            
            return True, "验证成功"
        except Exception as e:
            return False, f"验证失败：{str(e)}"

    def validate_headers(self, file_path):
        """验证文件的表头结构"""
        try:
            # 尝试从第4行和第5行读取表头
            errors = []
            for header_row in [3, 4]:  # 因为pandas的header参数是从0开始计数的
                try:
                    df = pd.read_excel(file_path, header=header_row)
                    is_valid, message = self.validate_headers_df(df)
                    if is_valid:
                        return True, "验证成功"
                    errors.append(f"第{header_row+1}行作为表头时：{message}")
                except Exception as e:
                    errors.append(f"读取第{header_row+1}行失败：{str(e)}")
            
            error_message = "表头验证失败：\n"
            error_message += "\n".join(errors)
            return False, error_message
        except Exception as e:
            return False, f"文件读取失败：{str(e)}"

    def process_file(self, file_path):
        """处理单个Excel文件"""
        try:
            # 获取Excel文件中的所有表格
            xl = pd.ExcelFile(file_path)
            sheet_names = xl.sheet_names

            # 依次验证每个表格
            all_errors = []
            for sheet_name in sheet_names:
                sheet_errors = []
                try:
                    # 尝试从第4行和第5行读取表头，使用engine='openpyxl'并忽略公式错误
                    for header_row in [3, 4]:  # 因为pandas的header参数是从0开始计数的
                        try:
                            # 读取当前表格，添加额外的参数来处理错误
                            df = pd.read_excel(
                                file_path,
                                sheet_name=sheet_name,
                                header=header_row,
                                engine='openpyxl',
                                keep_default_na=True,
                                na_values=[''],
                                dtype=str  # 将所有列都作为字符串读取
                            )
                            
                            # 验证表头
                            is_valid, message = self.validate_headers_df(df)
                            if is_valid:
                                # 删除空行（除序号外所有列都为空的行）
                                df = df.loc[~((df.iloc[:, 1:].isna().all(axis=1)) & (df.iloc[:, 0].notna()))]
                                
                                # 删除完全空白的行
                                df = df.dropna(how='all')
                                
                                # 处理时间格式
                                try:
                                    df['工作开始时间'] = pd.to_datetime(df['工作开始时间'], errors='coerce')
                                    df['工作结束时间'] = pd.to_datetime(df['工作结束时间'], errors='coerce')
                                except Exception as e:
                                    return None, f"时间格式转换失败：{str(e)}"
                                
                                # 重置索引
                                df = df.reset_index(drop=True)
                                
                                return df, f"使用表格：{sheet_name}"
                            else:
                                sheet_errors.append(f"第{header_row+1}行作为表头时：{message}")
                        except Exception as e:
                            sheet_errors.append(f"读取第{header_row+1}行失败，请检查Excel文件格式是否正确，确保没有合并单元格或特殊格式：{str(e)}")
                except Exception as e:
                    sheet_errors.append(f"处理失败：{str(e)}")
                
                if sheet_errors:
                    all_errors.append(f"\n表格 '{sheet_name}' 验证结果：\n" + "\n".join(sheet_errors))

            # 如果所有表格都验证失败
            error_message = f"文件 {os.path.basename(file_path)} 中没有找到有效的表格结构：\n"
            error_message += "\n".join(all_errors)
            return None, error_message

        except Exception as e:
            return None, f"处理失败：{str(e)}"

    def merge_files(self, file_paths):
        """合并多个Excel文件"""
        all_data = []
        all_errors = []
        
        for file_path in file_paths:
            try:
                # 首先验证文件是否存在
                if not os.path.exists(file_path):
                    all_errors.append(f"文件不存在：{file_path}")
                    continue
                    
                # 验证文件是否可以打开
                try:
                    xl = pd.ExcelFile(file_path)
                except Exception as e:
                    all_errors.append(f"无法打开文件 {os.path.basename(file_path)}：{str(e)}")
                    continue
                
                # 处理文件
                df, message = self.process_file(file_path)
                if df is not None:
                    # 验证数据有效性
                    if len(df) == 0:
                        all_errors.append(f"文件 {os.path.basename(file_path)} 没有有效数据")
                        continue
                    
                    # 验证必要列的数据类型
                    try:
                        # 验证时间格式
                        if pd.isna(df['工作开始时间']).all() or pd.isna(df['工作结束时间']).all():
                            all_errors.append(f"文件 {os.path.basename(file_path)} 的时间列全为空")
                            continue
                            
                        all_data.append(df)
                    except Exception as e:
                        all_errors.append(f"文件 {os.path.basename(file_path)} 数据验证失败：{str(e)}")
                        continue
                else:
                    all_errors.append(message)
            except Exception as e:
                all_errors.append(f"处理文件 {os.path.basename(file_path)} 时出错：{str(e)}")
        
        # 如果没有有效数据
        if not all_data:
            error_message = "合并失败，没有有效数据可合并。\n\n详细错误信息：\n"
            error_message += "\n".join(all_errors)
            return None, error_message
        
        try:
            # 合并所有数据
            self.merged_data = pd.concat(all_data, ignore_index=True)
            
            # 验证合并后的数据
            if len(self.merged_data) == 0:
                return None, "合并后的数据为空"
            
            # 删除只有序号列有内容的行
            self.merged_data = self.merged_data.loc[~((self.merged_data.iloc[:, 1:].isna().all(axis=1)) & (self.merged_data.iloc[:, 0].notna()))]
            
            # 验证是否还有数据
            if len(self.merged_data) == 0:
                return None, "删除无效行后没有剩余数据"
            
            # 按工作开始时间排序
            try:
                self.merged_data = self.merged_data.sort_values('工作开始时间')
            except Exception as e:
                return None, f"排序失败：{str(e)}"
            
            # 重置索引
            self.merged_data = self.merged_data.reset_index(drop=True)
            
            # 检查重复行
            self.check_duplicates()
            
            # 如果有错误但仍有可合并的数据，返回警告信息
            if all_errors:
                return self.merged_data, f"部分文件合并成功，但存在以下问题：\n" + "\n".join(all_errors)
            
            return self.merged_data, "合并成功"
            
        except Exception as e:
            error_message = f"合并数据时出错：{str(e)}\n\n此前的错误信息：\n"
            error_message += "\n".join(all_errors)
            return None, error_message

    def check_duplicates(self):
        """检查并标记重复行"""
        if self.merged_data is None:
            return
        
        # 找出重复行
        duplicates = self.merged_data[self.merged_data.duplicated(keep='first')]
        self.duplicate_rows = duplicates.index.tolist()

    def save_output(self, template_path, merged_data, output_path):
        """保存处理后的文件到模板"""
        if merged_data is None:
            return False, "没有数据可保存"
        
        try:
            # 加载模板文件
            wb = load_workbook(template_path)
            ws = wb.active
            
            # 保存原始列宽
            original_column_widths = {}
            for column in ws.columns:
                col_letter = column[0].column_letter
                original_column_widths[col_letter] = ws.column_dimensions[col_letter].width
            
            # 获取数据行数
            data_rows = len(merged_data)
            
            # 从第7行开始写入数据
            for row_idx, row_data in enumerate(merged_data.itertuples(), start=7):
                max_text_lines = 1  # 记录当前行中最大的文本行数
                
                for col_idx, value in enumerate(row_data[1:], start=1):  # 跳过索引列
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    # 处理序号列
                    if col_idx == 1:  # 序号列
                        cell.value = row_idx - 6  # 从1开始递增
                    # 处理时间格式
                    elif col_idx in [7, 8]:  # 工作开始时间和工作结束时间列
                        if pd.notna(value):  # 检查是否为空
                            cell.value = value.strftime("%Y-%m-%d")
                    else:
                        cell.value = value
                    
                    # 设置字体
                    cell.font = Font(name='宋体', size=9)
                    
                    # 设置对齐方式
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # 设置边框
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # 计算文本行数
                    if value and isinstance(value, str):
                        # 获取列宽
                        col_letter = cell.column_letter
                        col_width = original_column_widths.get(col_letter, 10)  # 默认宽度为10
                        
                        # 估算每行可以容纳的字符数（假设中文字符宽度为2，英文字符宽度为1）
                        chars_per_line = int(col_width / 2)  # 保守估计
                        if chars_per_line < 1:
                            chars_per_line = 1
                        
                        # 计算文本行数
                        text_length = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in str(value))
                        lines = (text_length + chars_per_line - 1) // chars_per_line
                        max_text_lines = max(max_text_lines, lines)
            
                # 设置行高（每行文字高度为8个单位，额外加10个单位作为边距）
                row_height = max(40, max_text_lines * 6 + 5)
                # 添加行高上限限制
                row_height = min(row_height, 180)  # 新增行高上限
                ws.row_dimensions[row_idx].height = row_height
            
            # 标记重复行
            light_red = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            for row_idx in self.duplicate_rows:
                for col in range(1, len(self.REQUIRED_COLUMNS) + 1):
                    cell = ws.cell(row=row_idx + 7, column=col)  # +7 因为数据从第7行开始
                    cell.fill = light_red
            
            # 恢复原始列宽
            for col_letter, width in original_column_widths.items():
                if width is not None:  # 只恢复有设置过宽度的列
                    ws.column_dimensions[col_letter].width = width
            
            # 保存为新文件
            wb.save(output_path)
            return True, "保存成功"
        except Exception as e:
            return False, f"保存失败：{str(e)}"

    def process_files(self, files):
        """处理多个Excel文件"""
        try:
            # 读取所有文件
            all_data = []
            for file in files:
                df = pd.read_excel(file, header=4)  # 从第5行开始读取
                if not self.validate_columns(df):
                    return None, f"文件 {os.path.basename(file)} 的列结构不正确"
                all_data.append(df)
            
            # 合并数据
            if not all_data:
                return None, "没有有效的数据可以合并"
            
            merged_df = pd.concat(all_data, ignore_index=True)
            
            # 删除只有序号列有内容的行
            merged_df = merged_df.loc[~((merged_df.iloc[:, 1:].isna().all(axis=1)) & (merged_df.iloc[:, 0].notna()))]
            
            # 按工作开始时间排序
            merged_df = merged_df.sort_values('工作开始时间')
            
            # 重置索引
            merged_df = merged_df.reset_index(drop=True)
            
            # 检查重复行
            self.duplicate_rows = []
            for idx, row in merged_df.iterrows():
                # 检查是否与之前的行重复
                if idx > 0:
                    prev_row = merged_df.iloc[idx-1]
                    if all(row[col] == prev_row[col] for col in self.REQUIRED_COLUMNS[1:]):  # 跳过序号列
                        self.duplicate_rows.append(idx)
            
            return merged_df, None
        except Exception as e:
            return None, f"处理文件时出错：{str(e)}" 