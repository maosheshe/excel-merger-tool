import sys
import os
from datetime import datetime
from PySide6.QtWidgets import (QApplication, QMainWindow, QPushButton, QVBoxLayout, 
                           QWidget, QFileDialog, QMessageBox, QListWidget, QLabel,
                           QHBoxLayout, QTableWidget, QTableWidgetItem, QProgressBar, QTextEdit)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QIcon, QColor
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from excel_processor import ExcelProcessor
from file_preview import FilePreviewWindow

class MergeWorker(QThread):
    progress_updated = Signal(int)  # 进度信号
    finished = Signal(bool, str)    # 完成信号
    error = Signal(str)             # 错误信号

    def __init__(self, processor, files, template_file):
        super().__init__()
        self.processor = processor
        self.files = files
        self.template_file = template_file

    def run(self):
        try:
            # 更新进度：开始处理
            self.progress_updated.emit(10)

            # 合并文件
            merged_data, message = self.processor.merge_files(self.files)
            if merged_data is None:
                self.error.emit(message)
                return

            # 更新进度：文件合并完成
            self.progress_updated.emit(50)

            # 生成输出文件名
            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
            current_date = datetime.now().strftime('%Y%m%d')
            
            # 从第一个文件中获取A2/A3内容
            try:
                wb = load_workbook(self.files[0])
                ws = wb.active
                a2_content = ws['A2'].value
                a3_content = ws['A3'].value
                content = a2_content if a2_content else a3_content
                if content:
                    output_file = os.path.join(desktop_path, f'附录2：{content}.xlsx')
                else:
                    output_file = os.path.join(desktop_path, f'附录2：营销现场作业计划审批表_{current_date}.xlsx')
            except Exception:
                output_file = os.path.join(desktop_path, f'附录2：营销现场作业计划审批表_{current_date}.xlsx')

            # 保存文件
            success, message = self.processor.save_output(self.template_file, merged_data, output_file)

            # 更新进度：完成
            self.progress_updated.emit(100)

            # 发送完成信号
            self.finished.emit(success, message if success else f"保存失败：{message}")

        except Exception as e:
            self.error.emit(str(e))

class ExcelMergerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.version = "v1.0.0"  # 添加版本号
        self.setWindowTitle(f"Excel文件合并工具 {self.version}")
        self.setMinimumSize(600, 400)
        
        # 获取程序运行目录
        self.exe_dir = os.path.dirname(os.path.abspath(sys.executable))
        if getattr(sys, 'frozen', False):
            # 如果是打包后的exe
            self.default_template = os.path.join(self.exe_dir, "输出模版.xlsx")
        else:
            # 如果是开发环境
            self.default_template = os.path.join(os.path.dirname(os.path.abspath(__file__)), "输出模版.xlsx")
        
        # 初始化变量
        self.selected_files = []
        self.template_file = None
        self.preview_window = None
        self.processor = ExcelProcessor()
        
        # 检查默认模板是否存在
        if os.path.exists(self.default_template):
            self.template_file = self.default_template
        
        self.init_ui()
        self.center_window()

    def center_window(self):
        """将窗口居中显示"""
        screen = QApplication.primaryScreen().geometry()
        size = self.geometry()
        x = (screen.width() - size.width()) // 2
        y = (screen.height() - size.height()) // 2
        self.move(x, y)

    def init_ui(self):
        """初始化UI"""
        # 创建主窗口部件
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        # 创建主布局
        layout = QVBoxLayout()
        
        # 创建按钮布局
        button_layout = QHBoxLayout()
        
        # 创建选择文件按钮
        self.select_button = QPushButton("选择Excel文件")
        self.select_button.clicked.connect(self.select_files)
        button_layout.addWidget(self.select_button)
        
        # 创建清除选择按钮
        self.clear_button = QPushButton("清除选择")
        self.clear_button.clicked.connect(self.clear_selection)
        button_layout.addWidget(self.clear_button)
        
        # 创建规范检查按钮
        self.preview_button = QPushButton("规范检查")
        self.preview_button.clicked.connect(self.preview_file)
        button_layout.addWidget(self.preview_button)
        
        layout.addLayout(button_layout)
        
        # 创建文件列表
        self.file_list = QListWidget()
        layout.addWidget(self.file_list)
        
        # 创建选择模板按钮
        self.template_button = QPushButton("选择输出模板")
        self.template_button.clicked.connect(self.select_template)
        layout.addWidget(self.template_button)
        
        # 创建模板标签
        self.template_label = QLabel("未选择模板文件" if not self.template_file else f"已选择模板：\n{os.path.basename(self.template_file)}")
        layout.addWidget(self.template_label)
        
        # 创建合并按钮
        self.merge_button = QPushButton("合并文件")
        self.merge_button.clicked.connect(self.merge_files)
        layout.addWidget(self.merge_button)
        
        # 创建进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # 创建状态标签
        self.status_label = QLabel("")
        layout.addWidget(self.status_label)
        
        main_widget.setLayout(layout)

    def clear_selection(self):
        """清除已选择的文件"""
        self.selected_files = []
        self.file_list.clear()
        self.status_label.setText('已清除选择的文件')

    def select_files(self):
        """选择Excel文件"""
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "选择Excel文件",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if files:
            # 检查总文件数是否超过限制
            total_files = len(self.selected_files) + len(files)
            if total_files > 5:
                QMessageBox.warning(self, '警告', '最多只能选择5个文件！')
                # 只添加能够添加的文件数量
                files = files[:5 - len(self.selected_files)]
            
            # 添加新选择的文件
            self.selected_files.extend(files)
            # 清空列表并重新显示所有文件
            self.file_list.clear()
            self.file_list.addItems([os.path.basename(f) for f in self.selected_files])
            self.status_label.setText(f'已选择 {len(self.selected_files)} 个文件')

    def select_template(self):
        """选择模板文件"""
        file, _ = QFileDialog.getOpenFileName(
            self,
            "选择输出模板",
            os.path.dirname(self.template_file) if self.template_file else self.exe_dir,
            "Excel Files (*.xlsx)"
        )
        
        if file:
            self.template_file = file
            self.template_label.setText(f"已选择模板：\n{os.path.basename(file)}")
            self.status_label.setText("已选择模板文件")

    def merge_files(self):
        """合并文件"""
        if not self.selected_files:
            QMessageBox.warning(self, "警告", "请先选择要合并的文件！")
            return
            
        if not self.template_file:
            QMessageBox.warning(self, "警告", "请先选择输出模板文件！")
            return
            
        # 显示进度条
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        # 创建处理线程
        self.worker = MergeWorker(self.processor, self.selected_files, self.template_file)
        
        # 连接信号
        self.worker.progress_updated.connect(self.progress_bar.setValue)
        self.worker.finished.connect(self.handle_merge_finished)
        self.worker.error.connect(self.handle_merge_error)
        
        # 开始处理
        self.worker.start()
        
        # 禁用按钮
        self.merge_button.setEnabled(False)
        self.select_button.setEnabled(False)
        self.template_button.setEnabled(False)
        
    def handle_merge_error(self, error_message):
        """处理合并错误"""
        self.progress_bar.setVisible(False)
        
        # 创建详细的错误信息对话框
        error_dialog = QMessageBox(self)
        error_dialog.setIcon(QMessageBox.Critical)
        error_dialog.setWindowTitle("错误")
        error_dialog.setText("合并过程中出现错误")
        
        # 分析错误信息
        if "详细错误信息：" in error_message:
            main_error, detailed_error = error_message.split("详细错误信息：", 1)
            error_dialog.setInformativeText(main_error.strip())
            error_dialog.setDetailedText(detailed_error.strip())
        else:
            error_dialog.setDetailedText(error_message)
        
        error_dialog.setStandardButtons(QMessageBox.Ok)
        
        # 调整对话框大小
        text_browser = error_dialog.findChild(QTextEdit)
        if text_browser is not None:
            text_browser.setMinimumSize(600, 400)
        
        error_dialog.exec_()
        
        # 重新启用按钮
        self.merge_button.setEnabled(True)
        self.select_button.setEnabled(True)
        self.template_button.setEnabled(True)
        
        # 更新状态
        self.status_label.setText("合并失败，请查看错误信息")
        
    def handle_merge_finished(self, success, message):
        """处理合并完成"""
        self.progress_bar.setVisible(False)
        
        if success:
            # 检查是否有警告信息
            if "但存在以下问题：" in message:
                warning_dialog = QMessageBox(self)
                warning_dialog.setIcon(QMessageBox.Warning)
                warning_dialog.setWindowTitle("部分成功")
                warning_dialog.setText("文件已合并，但存在一些问题")
                warning_dialog.setDetailedText(message)
                warning_dialog.setStandardButtons(QMessageBox.Ok)
                
                # 调整对话框大小
                text_browser = warning_dialog.findChild(QTextEdit)
                if text_browser is not None:
                    text_browser.setMinimumSize(600, 400)
                
                warning_dialog.exec_()
                self.status_label.setText("合并完成，但有部分问题")
            else:
                QMessageBox.information(self, "成功", message)
                self.status_label.setText("合并完成！")
        else:
            # 创建详细的错误信息对话框
            error_dialog = QMessageBox(self)
            error_dialog.setIcon(QMessageBox.Critical)
            error_dialog.setWindowTitle("错误")
            error_dialog.setText("合并失败")
            
            # 分析错误信息
            if "详细错误信息：" in message:
                main_error, detailed_error = message.split("详细错误信息：", 1)
                error_dialog.setInformativeText(main_error.strip())
                error_dialog.setDetailedText(detailed_error.strip())
            else:
                error_dialog.setDetailedText(message)
            
            error_dialog.setStandardButtons(QMessageBox.Ok)
            
            # 调整对话框大小
            text_browser = error_dialog.findChild(QTextEdit)
            if text_browser is not None:
                text_browser.setMinimumSize(600, 400)
            
            error_dialog.exec_()
            self.status_label.setText("合并失败，请查看错误信息")
        
        # 重新启用按钮
        self.merge_button.setEnabled(True)
        self.select_button.setEnabled(True)
        self.template_button.setEnabled(True)

    def preview_file(self):
        """打开文件预览窗口"""
        if not self.preview_window:
            self.preview_window = FilePreviewWindow()
        self.preview_window.show()

def main():
    app = QApplication(sys.argv)
    window = ExcelMergerApp()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main() 